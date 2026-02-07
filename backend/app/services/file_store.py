from __future__ import annotations

import json
import os
import re
import shutil
from datetime import date, datetime, timezone
from numbers import Number
from pathlib import Path
from typing import Any
from uuid import uuid4

import pandas as pd
from fastapi import UploadFile
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ALLOWED_SUFFIXES = {".csv", ".xlsx"}
LAYOUT_TEXT_KEYWORDS = {"telefono", "phone", "cell", "cf", "codice", "cap", "piva", "iban"}
LAYOUT_DATE_KEYWORDS = {"data", "date"}
LAYOUT_AMOUNT_KEYWORDS = {"importo", "totale", "amount", "€"}


def _layout_pack_enabled() -> bool:
    raw = os.getenv("LAYOUT_PACK", "1").strip().lower()
    return raw not in {"0", "false", "off", "no"}


def _utcnow_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _safe_name(filename: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]", "_", filename)


def _is_non_empty(value: Any) -> bool:
    return value is not None and value != ""


def _is_numeric_value(value: Any) -> bool:
    return isinstance(value, Number) and not isinstance(value, bool)


def _is_date_value(value: Any) -> bool:
    return isinstance(value, (date, datetime))


def _header_bucket(header_value: Any) -> tuple[bool, bool, bool]:
    text = str(header_value or "").strip().lower()
    is_text = any(keyword in text for keyword in LAYOUT_TEXT_KEYWORDS)
    is_date = any(keyword in text for keyword in LAYOUT_DATE_KEYWORDS)
    is_amount = any(keyword in text for keyword in LAYOUT_AMOUNT_KEYWORDS)
    return is_text, is_date, is_amount


def _last_non_empty_row(ws, max_col: int) -> int:
    for row_idx in range(ws.max_row, 0, -1):
        for col_idx in range(1, max_col + 1):
            if _is_non_empty(ws.cell(row=row_idx, column=col_idx).value):
                return row_idx
    return 1


def _last_non_empty_col(ws, max_row: int) -> int:
    for col_idx in range(ws.max_column, 0, -1):
        for row_idx in range(1, max_row + 1):
            if _is_non_empty(ws.cell(row=row_idx, column=col_idx).value):
                return col_idx
    return 1


def _table_name(ws, base_name: str = "ResultTable") -> str:
    existing = set(ws.tables.keys())
    if base_name not in existing:
        return base_name
    counter = 2
    while f"{base_name}{counter}" in existing:
        counter += 1
    return f"{base_name}{counter}"


def apply_layout_pack(path: Path) -> None:
    workbook = load_workbook(path)
    worksheet = workbook.active

    if worksheet.max_column <= 0:
        workbook.save(path)
        return

    last_row = _last_non_empty_row(worksheet, worksheet.max_column)
    last_col = _last_non_empty_col(worksheet, last_row)
    if last_col <= 0:
        workbook.save(path)
        return

    table_ref = f"A1:{get_column_letter(last_col)}{max(last_row, 1)}"

    if last_row >= 2:
        table = Table(displayName=_table_name(worksheet), ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    worksheet.auto_filter.ref = table_ref
    worksheet.freeze_panes = "A2"

    header_font = Font(bold=True)
    for col_idx in range(1, last_col + 1):
        header_cell = worksheet.cell(row=1, column=col_idx)
        header_cell.font = header_font
        header_cell.alignment = Alignment(
            vertical="center",
            horizontal=header_cell.alignment.horizontal,
            wrap_text=True,
        )
    worksheet.row_dimensions[1].height = 24

    for col_idx in range(1, last_col + 1):
        header_value = worksheet.cell(row=1, column=col_idx).value
        is_text_col, is_date_col, is_amount_col = _header_bucket(header_value)

        max_len = len(str(header_value or ""))
        non_empty_cells = 0
        numeric_cells = 0
        for row_idx in range(2, last_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if not _is_non_empty(cell_value):
                continue
            non_empty_cells += 1
            max_len = max(max_len, len(str(cell_value)))
            if _is_numeric_value(cell_value):
                numeric_cells += 1

        numeric_column = (
            not is_text_col
            and non_empty_cells > 0
            and numeric_cells == non_empty_cells
        )

        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].width = max(10, min(45, max_len + 2))

        for row_idx in range(2, last_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell_value = cell.value
            if not _is_non_empty(cell_value):
                continue

            cell.alignment = Alignment(
                vertical="center",
                horizontal="right" if numeric_column and _is_numeric_value(cell_value) else cell.alignment.horizontal,
                wrap_text=cell.alignment.wrap_text,
            )

            if is_text_col:
                cell.number_format = "@"
            elif is_date_col and _is_date_value(cell_value):
                cell.number_format = "DD/MM/YYYY"
            elif is_amount_col and _is_numeric_value(cell_value):
                cell.number_format = "€ #,##0.00"

    workbook.save(path)


class FileStore:
    def __init__(self, data_dir: Path) -> None:
        self.data_dir = data_dir
        self.upload_dir = self.data_dir / "uploads"
        self.result_dir = self.data_dir / "results"
        self.upload_meta_dir = self.data_dir / "meta" / "uploads"
        self.result_meta_dir = self.data_dir / "meta" / "results"
        for directory in (
            self.upload_dir,
            self.result_dir,
            self.upload_meta_dir,
            self.result_meta_dir,
        ):
            directory.mkdir(parents=True, exist_ok=True)

    def save_upload(self, upload: UploadFile) -> dict[str, Any]:
        filename = upload.filename or "dataset.csv"
        suffix = Path(filename).suffix.lower()
        if suffix not in ALLOWED_SUFFIXES:
            raise ValueError("Unsupported file type. Use .csv or .xlsx.")

        file_id = uuid4().hex
        stored_filename = f"{file_id}{suffix}"
        stored_path = self.upload_dir / stored_filename

        with stored_path.open("wb") as destination:
            shutil.copyfileobj(upload.file, destination)

        metadata = {
            "file_id": file_id,
            "filename": _safe_name(filename),
            "suffix": suffix,
            "stored_path": str(stored_path),
            "file_size_bytes": int(stored_path.stat().st_size),
            "created_at": _utcnow_iso(),
        }
        self._write_json(self.upload_meta_dir / f"{file_id}.json", metadata)
        return metadata

    def get_upload_meta(self, file_id: str) -> dict[str, Any]:
        return self._read_json(self.upload_meta_dir / f"{file_id}.json")

    def get_result_meta(self, result_id: str) -> dict[str, Any]:
        return self._read_json(self.result_meta_dir / f"{result_id}.json")

    def load_upload_df(self, file_id: str) -> pd.DataFrame:
        meta = self.get_upload_meta(file_id)
        path = Path(meta["stored_path"])
        if not path.exists():
            raise FileNotFoundError(f"Upload not found for file_id={file_id}")

        suffix = meta["suffix"]
        if suffix == ".csv":
            return pd.read_csv(path)
        if suffix == ".xlsx":
            return pd.read_excel(path, engine="openpyxl")
        raise ValueError("Unsupported file format in metadata.")

    def save_result(
        self,
        df: pd.DataFrame,
        source_file_id: str,
        output_format: str,
    ) -> dict[str, Any]:
        if output_format not in {"csv", "xlsx"}:
            raise ValueError("output_format must be csv or xlsx")

        result_id = uuid4().hex
        suffix = f".{output_format}"
        result_path = self.result_dir / f"{result_id}{suffix}"

        if output_format == "csv":
            df.to_csv(result_path, index=False)
        else:
            df.to_excel(result_path, index=False)
            if _layout_pack_enabled():
                apply_layout_pack(result_path)

        metadata = {
            "result_id": result_id,
            "source_file_id": source_file_id,
            "output_format": output_format,
            "stored_path": str(result_path),
            "created_at": _utcnow_iso(),
        }
        self._write_json(self.result_meta_dir / f"{result_id}.json", metadata)
        return metadata

    @staticmethod
    def _read_json(path: Path) -> dict[str, Any]:
        if not path.exists():
            raise FileNotFoundError(path.name)
        return json.loads(path.read_text(encoding="utf-8"))

    @staticmethod
    def _write_json(path: Path, payload: dict[str, Any]) -> None:
        path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
