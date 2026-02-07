from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from app.services.file_store import FileStore


def test_save_result_applies_layout_pack_for_xlsx(tmp_path) -> None:
    store = FileStore(tmp_path / "data")
    source = pd.DataFrame(
        {
            "telefono": ["0012345678", "0099988877"],
            "data_operazione": [pd.Timestamp("2026-02-01"), pd.Timestamp("2026-03-05")],
            "importo": [12.5, 999.0],
            "cliente": ["Anna Rossi", "Mario Bianchi"],
        }
    )

    metadata = store.save_result(source, source_file_id="source-1", output_format="xlsx")
    workbook = load_workbook(metadata["stored_path"])
    worksheet = workbook.active

    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref is not None
    assert len(worksheet.tables) == 1
    assert next(iter(worksheet.tables.values())).tableStyleInfo.name == "TableStyleMedium9"

    assert worksheet["A2"].number_format == "@"
    assert worksheet["B2"].number_format == "DD/MM/YYYY"
    assert worksheet["C2"].number_format == "€ #,##0.00"

    assert worksheet.column_dimensions["A"].width >= 10
    assert worksheet.column_dimensions["A"].width <= 45


def test_save_result_skips_layout_pack_when_disabled(tmp_path, monkeypatch) -> None:
    monkeypatch.setenv("LAYOUT_PACK", "0")
    store = FileStore(tmp_path / "data")
    source = pd.DataFrame({"colonna": [1, 2, 3]})

    metadata = store.save_result(source, source_file_id="source-2", output_format="xlsx")
    workbook = load_workbook(metadata["stored_path"])
    worksheet = workbook.active

    assert worksheet.freeze_panes is None
    assert len(worksheet.tables) == 0
